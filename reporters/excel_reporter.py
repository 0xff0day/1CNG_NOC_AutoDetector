"""
Excel Report Generator

Generates Excel reports from monitoring data.
Uses openpyxl for formatting, charts, and styling.
"""

from __future__ import annotations

from typing import List, Dict, Optional, Any
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


@dataclass
class ReportColumn:
    """Excel column definition."""
    header: str
    field: str
    width: int = 15
    format: str = "general"


class ExcelReportGenerator:
    """
    Generates formatted Excel reports.
    
    Features:
    - Multiple worksheets
    - Charts and graphs
    - Conditional formatting
    - Auto-filter and sorting
    """
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_device_report(
        self,
        device_data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        Generate device status report.
        
        Args:
            device_data: List of device metric dicts
            filename: Output filename (auto-generated if None)
        
        Returns:
            Path to generated file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            logger.error("openpyxl not installed")
            raise
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Device Status"
        
        # Define columns
        columns = [
            ReportColumn("Device ID", "device_id", 20),
            ReportColumn("Hostname", "hostname", 25),
            ReportColumn("Type", "device_type", 15),
            ReportColumn("Status", "status", 12),
            ReportColumn("Health Score", "health_score", 12, "number"),
            ReportColumn("CPU %", "cpu_usage", 10, "percent"),
            ReportColumn("Memory %", "memory_usage", 10, "percent"),
            ReportColumn("Disk %", "disk_usage", 10, "percent"),
            ReportColumn("Last Seen", "last_seen", 20, "datetime"),
            ReportColumn("Alerts", "alert_count", 10, "number"),
        ]
        
        # Header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = col.width
        
        # Data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, device in enumerate(device_data, 2):
            for col_idx, col in enumerate(columns, 1):
                value = device.get(col.field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Conditional formatting for status
                if col.field == "status":
                    if value == "healthy":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif value == "warning":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700")
                    elif value == "critical":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                
                # Format percentages
                if col.format == "percent" and isinstance(value, (int, float)):
                    cell.number_format = '0.0"%"'
        
        # Add auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(device_data)+1}"
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add summary sheet
        self._add_summary_sheet(wb, device_data)
        
        # Save
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"device_report_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath}")
        
        return str(filepath)
    
    def _add_summary_sheet(self, wb, device_data: List[Dict]) -> None:
        """Add summary statistics sheet."""
        from openpyxl import Workbook
        from openpyxl.chart import PieChart, Reference
        
        ws = wb.create_sheet("Summary")
        
        # Calculate statistics
        total = len(device_data)
        healthy = sum(1 for d in device_data if d.get("status") == "healthy")
        warning = sum(1 for d in device_data if d.get("status") == "warning")
        critical = sum(1 for d in device_data if d.get("status") == "critical")
        
        avg_health = sum(d.get("health_score", 0) for d in device_data) / total if total > 0 else 0
        
        # Summary data
        ws["A1"] = "Metric"
        ws["B1"] = "Value"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        summary_data = [
            ("Total Devices", total),
            ("Healthy", healthy),
            ("Warning", warning),
            ("Critical", critical),
            ("Average Health Score", round(avg_health, 1)),
        ]
        
        for row_idx, (metric, value) in enumerate(summary_data, 2):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Add pie chart
        chart = PieChart()
        chart.title = "Device Status Distribution"
        
        labels = Reference(ws, min_col=1, min_row=3, max_row=5)
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        ws.add_chart(chart, "D2")
    
    def generate_alert_report(
        self,
        alerts: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        Generate alert history report.
        
        Args:
            alerts: List of alert dicts
            filename: Output filename
        
        Returns:
            Path to generated file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed")
            raise
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Alerts"
        
        headers = ["ID", "Timestamp", "Device", "Severity", "Variable", "Message", "Status"]
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row_idx, alert in enumerate(alerts, 2):
            ws.cell(row=row_idx, column=1, value=alert.get("id"))
            ws.cell(row=row_idx, column=2, value=alert.get("timestamp"))
            ws.cell(row=row_idx, column=3, value=alert.get("device_id"))
            ws.cell(row=row_idx, column=4, value=alert.get("severity"))
            ws.cell(row=row_idx, column=5, value=alert.get("variable"))
            ws.cell(row=row_idx, column=6, value=alert.get("message"))
            ws.cell(row=row_idx, column=7, value=alert.get("status", "active"))
            
            # Severity coloring
            severity = alert.get("severity", "").lower()
            if severity in ["critical", "emergency"]:
                ws.cell(row=row_idx, column=4).fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                ws.cell(row=row_idx, column=4).font = Font(color="FFFFFF", bold=True)
        
        # Auto-filter
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(alerts)+1}"
        
        # Save
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"alert_report_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
